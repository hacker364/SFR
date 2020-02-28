import os
import time
import cv2
from openpyxl import Workbook
from openpyxl import load_workbook

def preview_image(image, name="window", time=100, resize=True):
    if resize:
        cv2.imshow(name, cv2.resize(image, (400, 400)))
    else:
        cv2.imshow(name, image)
    cv2.waitKey(time)

def take_image(name):
    os.chdir("/home/pi/SFR/trainData")
    if os.path.exists(name):
        print("Person with same Name exists")
        exit
    else:
        os.makedirs(name)
    
    video_capture = cv2.VideoCapture(0)
    success, image = video_capture.read()
    count = 30 
        
    while success:
        success, image = video_capture.read() 
        cv2.imwrite(name +"/frame%d.jpg" % count, image)     # save frame as PNG file
        count -= 1
        if count == 0:
            break
        preview_image(image)

if __name__ == "__main__":
    data = []
    os.chdir('/home/pi/SFR')
    wb = load_workbook(filename = 'base.xlsx')
    sheet = wb.active
    max_rows = len(sheet['A']) + 1
    for cells in range(2, max_rows):
        name = sheet["A" + str(cells)].value
        if(name == None):
            max_rows = cells
            break
    
    name = input("Enter the name of student: ")
    ID = input("Enter the ID of student: ")
    data.append(name)
    data.append(ID)

    for rows in data:
        sheet.cell(row=max_rows, column=1, value=name)
        sheet.cell(row=max_rows, column=2, value=ID)
        sheet.cell(row=max_rows, column=3, value=0)
        sheet.cell(row=max_rows, column=4, value=0)
    
    wb.save(filename = 'base.xlsx')
    take_image(name)
